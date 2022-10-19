import openpyxl

inventory_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inventory_file['Sheet1']
print(product_list)

products_per_supplier = {}
inventory_per_supplier = {}
inventory_less_than_ten = {}

for product_row in range(2,product_list.max_row+1):
    supplier_name = product_list.cell(product_row, 4).value # The value in fourth column
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    inventory_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5) # Don't add value as we need an empty cell
    
    print(supplier_name)
    
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1 
    
    if supplier_name in inventory_per_supplier:
        inventory_per_supplier[supplier_name] += (inventory * price)
    else:
        inventory_per_supplier[supplier_name] = (inventory * price)
        
    if inventory < 10:
        inventory_less_than_ten[inventory_number] = int(inventory)
        
    inventory_price.value = inventory * price
         
    

print(products_per_supplier) 
print(inventory_per_supplier) 
print(inventory_less_than_ten)

inventory_file.save('inventory_with_total_value.csv')